'''
将数据集转化为cora.cite格式，
cited_paper_id \t citing_paper_id
'''

from openpyxl import load_workbook, Workbook
from gensim import corpora
import os

def cited2citing(input, output, method):
    row = len(method)
    # # list 2 dic
    dic = corpora.Dictionary(method)
    dic.save_as_text('../data/dic.txt')
    dic_set = dic.token2id
    # values = []
    # for word in method:
    #     values.append(dic_set[word])

    final = []
    for i in range(row):
        for ins in input[i]:
            if ins == 'int' or ins=='long' or ins == 'float' or ins=='double' or ins=='boolean' or ins == "String" or ins == 'Object' or ins == 'byte':
                # print("skip")
                continue
            for j in range(row):
                # 如果输入在输出中出现
                if ins in output[j]:
                    cited_name = method[j][0]
                    citing_name = method[i][0]
                    cited_id = dic_set[cited_name]
                    citing_id = dic_set[citing_name]
                    final.append(str(cited_id)+'\t'+str(citing_id)+'\n')
        print(i/row)
    print("文件计算完毕")
    return final

def write_cited(data, filename):
    # 写之前，先检验文件是否存在，存在就删掉
    if os.path.exists(filename):
        os.remove(filename)

    # 以写的方式打开文件，如果文件不存在，就会自动创建
    file_write_obj = open(filename, 'w')
    for var in data:
        file_write_obj.writelines(var)
        # file_write_obj.write('\n')
    file_write_obj.close()









# # 创建一个worksheet
#     for k in j.split(" "):
#         if len(k) >= 2:
#             input.append(str(k))
# for p in
#     for q in p.split(" "):
#         if len(q) >= 2:
#             output.append(str(q))
# print(i)
# wb1 = Workbook()
# ws1 = wb1.active
# ws1.title = "sheet1"
# wb2 = Workbook()
# ws2 = wb2.active
# ws2.title = "sheet1"
# print("文件已创建")
# input_line = ""
# output_line = ""
#
# print(len(input))
# print(len(output))
# for j in range(len(input)):
#     for i in range(1, rows + 1):
#         if str(input[j]) in str(table.cell(row=i, column=2).value):
#             input_line += str(table.cell(row=i, column=1).value)+","
#     ws1.cell(row=j + 1, column=2).value = input_line[0:-1]
#     ws1.cell(row = j+1,column=1).value = input[j].replace('[','').replace(']','').replace('\'','')
#     input_line = ""
#     print(j)
#
# for k in range(len(output)):
#     for i in range(1, rows + 1):
#         if str(output[k]) in str(table.cell(row=i, column=3).value):
#             output_line += str(table.cell(row=i, column=1).value) + ","
#     ws2.cell(row=k + 1, column=2).value = output_line[0:-1]
#     ws2.cell(row = k+1,column=1).value = output[k].replace('[','').replace(']','').replace('\'','')
#     output_line = ""
#     print(k)
#
#
#
# print("文件写入完毕")
# # 保存
# wb1.save(filename="magic_input_4_test.xlsx")
# wb2.save(filename="magic_output_4_test.xlsx")
# print("文件已保存")
